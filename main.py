import os
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ---------------- 1) Load Data ---------------- #
def load_ipl_data(file_path: str):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"{file_path} does not exist.")
    df = pd.read_excel(file_path) if file_path.endswith(".xlsx") else pd.read_csv(file_path)
    return df

# ---------------- 2) Clean Data ---------------- #
def clean_ipl_data(df):
    df = df.fillna({
        'Player': 'Unknown',
        'Role': 'Unknown',
        'Team': 'Unknown',
        'Matches': 0,
        'Runs': 0,
        'Bat_SR': 0,
        'Wickets': 0,
        'Econ': 0
    })
    # Ensure numeric columns are numeric
    for col in ['Runs', 'Wickets', 'Bat_SR', 'Econ']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    df = df.drop_duplicates(subset=['Player', 'Team'])
    return df

# ---------------- 3) Analyze Data ---------------- #
def analyze_ipl(df):
    top_runs = df.nlargest(10, 'Runs')[['Player', 'Team', 'Runs']]
    top_wickets = df.nlargest(10, 'Wickets')[['Player', 'Team', 'Wickets']]
    
    # Summary stats
    summary = pd.Series({
        'Avg_Batting_SR': df['Bat_SR'].mean(),
        'Avg_Bowling_Econ': df['Econ'].replace(0, pd.NA).dropna().astype(float).mean()
    })
    
    return top_runs, top_wickets, summary

# ---------------- 4) Generate Charts ---------------- #
def generate_charts(df, charts_dir='charts'):
    os.makedirs(charts_dir, exist_ok=True)
    paths = {}

    # Runs Distribution
    plt.figure(figsize=(8,5))
    sns.histplot(df['Runs'], bins=20, kde=True)
    plt.title("Runs Distribution")
    runs_path = os.path.join(charts_dir, 'runs_dist.png')
    plt.savefig(runs_path, bbox_inches='tight')
    plt.close()
    paths['runs'] = runs_path

    # Wickets Distribution
    plt.figure(figsize=(8,5))
    sns.histplot(df['Wickets'], bins=20, kde=True)
    plt.title("Wickets Distribution")
    wickets_path = os.path.join(charts_dir, 'wickets_dist.png')
    plt.savefig(wickets_path, bbox_inches='tight')
    plt.close()
    paths['wickets'] = wickets_path

    # Team-wise Runs
    plt.figure(figsize=(10,5))
    team_runs = df.groupby('Team')['Runs'].sum().sort_values(ascending=False)
    sns.barplot(x=team_runs.index, y=team_runs.values)
    plt.xticks(rotation=45)
    plt.title("Total Runs by Team")
    team_runs_path = os.path.join(charts_dir, 'team_runs.png')
    plt.savefig(team_runs_path, bbox_inches='tight')
    plt.close()
    paths['team_runs'] = team_runs_path

    # Team-wise Wickets
    plt.figure(figsize=(10,5))
    team_wkts = df.groupby('Team')['Wickets'].sum().sort_values(ascending=False)
    sns.barplot(x=team_wkts.index, y=team_wkts.values)
    plt.xticks(rotation=45)
    plt.title("Total Wickets by Team")
    team_wkts_path = os.path.join(charts_dir, 'team_wickets.png')
    plt.savefig(team_wkts_path, bbox_inches='tight')
    plt.close()
    paths['team_wickets'] = team_wkts_path

    return paths

# ---------------- 5) Export Report ---------------- #
def export_report(output_file, top_runs, top_wickets, summary):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        top_runs.to_excel(writer, sheet_name="Top_Run_Scorers", index=False)
        top_wickets.to_excel(writer, sheet_name="Top_Wicket_Takers", index=False)
        summary_df = summary.rename("Value").reset_index().rename(columns={'index':'Metric'})
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

# ---------------- 6) Embed Charts in Excel ---------------- #
def embed_charts(excel_file, chart_paths):
    wb = load_workbook(excel_file)
    if 'Charts' not in wb.sheetnames:
        wb.create_sheet('Charts')
    ws = wb['Charts']
    row = 1
    for label, path in chart_paths.items():
        img = XLImage(path)
        ws.add_image(img, f"A{row}")
        row += 25  # leave enough space for next chart
    wb.save(excel_file)

# ---------------- Orchestration ---------------- #
def main():
    source_file = "ipl_data.xlsx"       # <-- Update with your actual file
    output_excel = "ipl_analysis_report.xlsx"
    charts_dir = "charts"

    df = load_ipl_data(source_file)

    df = clean_ipl_data(df)

    top_runs, top_wickets, summary = analyze_ipl(df)

    export_report(output_excel, top_runs, top_wickets, summary)

    chart_paths = generate_charts(df, charts_dir)
    embed_charts(output_excel, chart_paths)

    # Console Output
    print("✅ IPL Analysis Completed!\n")
    print("Top Run Scorers:\n", top_runs)
    print("\nTop Wicket Takers:\n", top_wickets)
    print("\nSummary Stats:\n", summary)
    print(f"\nReport saved to: {output_excel}")

if __name__ == "__main__":
    main()
